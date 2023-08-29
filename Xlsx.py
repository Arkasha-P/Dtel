from openpyxl import load_workbook
from openpyxl.styles import (Border, Alignment, Font, Side)
import math
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

file_path = "txt/users.xlsx"
sheet_name = "Лист1"



# Загрузка файла
workbook = load_workbook(file_path)

# Получение активного листа
sheet = workbook.active

# Подсчет заполненных строк
count = 0
for row in sheet.iter_rows():
    # Проверка, если значение в первой ячейке строки не равно None
    if row[0].value is not None:
        count += 1

print("Количество заполненных строк:", count)

for i, cell in enumerate(sheet['A'], start=1):
    value = cell.value
    sheet.cell(row=i*3, column=5).value = value

for i, cell in enumerate(sheet['B'], start=1):
    value = cell.value
    sheet.cell(row=i*3+1, column=5).value = value

for i, cell in enumerate(sheet['C'], start=1):
    value = cell.value
    sheet.cell(row=i*3+2, column=5).value = value

for i, cell in enumerate(sheet['A'], start=1):
    sheet.cell(row=i*3, column=4).value = "Квартира:"
    sheet.cell(row=i*3+1, column=4).value = "Логин:"
    sheet.cell(row=i*3+2, column=4).value = "Пароль:"
    if i == count: break

for i, cell in enumerate(sheet['A'], start=1):
    sheet.cell(row=i, column=1).value = ""
    sheet.cell(row=i, column=2).value = ""
    sheet.cell(row=i, column=3).value = ""
    if i == count: break

count = 0
for row in sheet.iter_rows():
    # Проверка, если значение в первой ячейке строки не равно None
    if row[3].value is not None:
        count += 1

print(count)

cell_range = f'd3:e{count+2}'
sheet.move_range(cell_range, rows=-2, cols=-3, translate=False)


count = 0
for row in sheet.iter_rows():
    # Проверка, если значение в первой ячейке строки не равно None
    if row[1].value is not None:
        count += 1


bd = Side(style='thin', color="000000")
i = 1
while i <= count:
    sheet[f'a{i}'].font = Font(bold=True, color='000000', name='Arial', size=11)
    sheet[f'b{i}'].font = Font(bold=True, color='000000', name='Arial', size=11)
    sheet[f'a{i}'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sheet[f'b{i}'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sheet[f'a{i}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'b{i}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[i].height = float(35)
    i += 1



width = 12
sheet.column_dimensions['A'].width = width
sheet.column_dimensions['B'].width = width
sheet.column_dimensions['C'].width = width
sheet.column_dimensions['D'].width = width
sheet.column_dimensions['E'].width = width
sheet.column_dimensions['F'].width = width
sheet.column_dimensions['G'].width = width
sheet.column_dimensions['H'].width = width
sheet.column_dimensions['I'].width = width
sheet.column_dimensions['J'].width = width
sheet.column_dimensions['K'].width = width
sheet.column_dimensions['L'].width = width
sheet.column_dimensions['M'].width = width
sheet.column_dimensions['N'].width = width
sheet.column_dimensions['O'].width = width
sheet.column_dimensions['P'].width = width
sheet.column_dimensions['Q'].width = width
sheet.column_dimensions['R'].width = width
sheet.column_dimensions['S'].width = width
sheet.column_dimensions['T'].width = width
sheet.column_dimensions['U'].width = width
sheet.column_dimensions['V'].width = width
sheet.column_dimensions['W'].width = width
sheet.column_dimensions['X'].width = width
sheet.column_dimensions['Y'].width = width
sheet.column_dimensions['Z'].width = width

### PAGE 1

cell_range = f'a22:b42'
sheet.move_range(cell_range, rows= -21, cols=2)

cell_range = f'a43:b63'
sheet.move_range(cell_range, rows= -42, cols=4)

### PAGE 2

cell_range = f'a64:b84'
sheet.move_range(cell_range, rows= -63, cols=7)

cell_range = f'a85:b105'
sheet.move_range(cell_range, rows= -84, cols=9)

cell_range = f'a106:b126'
sheet.move_range(cell_range, rows= -105, cols=11)

### PAGE 3

cell_range = f'a127:b147'
sheet.move_range(cell_range, rows= -126, cols=14)

cell_range = f'a148:b168'
sheet.move_range(cell_range, rows= -147, cols=16)

cell_range = f'a169:b189'
sheet.move_range(cell_range, rows= -168, cols=18)

### PAGE 4

cell_range = f'a190:b210'
sheet.move_range(cell_range, rows= -189, cols=21)

cell_range = f'a211:b231'
sheet.move_range(cell_range, rows= -210, cols=23)

cell_range = f'a232:b252'
sheet.move_range(cell_range, rows= -231, cols=25)


# columms = math.ceil(count / 21)
# print(columms)

# list2 = []
# for list1 in range(1,1000,20): list2.append(list1)

# Сохранение изменений
workbook.save(file_path)

